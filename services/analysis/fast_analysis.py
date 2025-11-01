# Excel-to-AI Implementation: Fast Estimates Backend Code

**File**: `services/analysis/fast_analysis.py`  
**Purpose**: Wire together PDF extraction, compliance checking, and estimate generation  
**Target Time**: 5-10 minutes from upload to deliverables

---

## Core Analysis Pipeline

```python
# fast_analysis.py - Main entry point for Excel-to-AI bridge

from fastapi import FastAPI, UploadFile, BackgroundTasks, WebSocket
from fastapi.responses import FileResponse
import asyncio
import aiohttp
from datetime import datetime, timezone
from pydantic import BaseModel
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import PatternFill, Font
import pypdf
import io
import json
from enum import Enum

app = FastAPI(title="Eagle Eye Fast Analysis")

# ============================================================================
# 1. DATA MODELS
# ============================================================================

class ProjectInput(BaseModel):
    """User input: project basics"""
    project_name: str
    client_name: str
    client_email: str
    address: str
    city: str
    state: str
    zip_code: str
    jurisdiction: str
    scope: str
    special_conditions: Optional[str] = None

class ComponentData(BaseModel):
    """Individual component (window, door, etc.)"""
    component_type: str  # "windows", "doors", "framing", etc.
    quantity: int
    unit: str  # "each", "sf", "bf", etc.
    size_spec: Optional[str] = None
    location: Optional[str] = None
    notes: Optional[str] = None
    source: str  # "excel" or "ai_extracted"
    confidence: float = 1.0  # 0.0-1.0 confidence if AI-extracted

class AnalysisResult(BaseModel):
    """Complete analysis result"""
    project_id: str
    project_info: ProjectInput
    components: List[ComponentData]
    findings: List[Dict]  # Code compliance findings
    estimate: Dict  # Cost estimate breakdown
    regional_factors: Dict
    analysis_timestamp: str
    status: str  # "complete", "processing", "error"

# ============================================================================
# 2. MAIN ANALYSIS ENDPOINT
# ============================================================================

@app.post("/api/v1/analyze")
async def analyze_project(
    project_id: str,
    excel_file: UploadFile = None,
    pdf_files: List[UploadFile] = None,
    background_tasks: BackgroundTasks = None
):
    """
    Main entry point: Accept Excel + PDFs, kick off analysis asynchronously.
    Returns immediately with status, processes in background.
    """
    
    if not project_id:
        project_id = f"proj_{datetime.now(timezone.utc).timestamp()}"
    
    # Read uploaded files into memory
    excel_data = None
    pdf_data = []
    
    if excel_file:
        excel_bytes = await excel_file.read()
        excel_data = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    
    if pdf_files:
        for pdf_file in pdf_files:
            pdf_bytes = await pdf_file.read()
            pdf_data.append(pdf_bytes)
    
    # Start analysis asynchronously
    background_tasks.add_task(
        run_full_analysis,
        project_id=project_id,
        excel_workbook=excel_data,
        pdf_bytes_list=pdf_data
    )
    
    return {
        "project_id": project_id,
        "status": "processing",
        "message": "Analysis started. You'll receive email when complete.",
        "estimated_minutes": 5
    }

@app.websocket("/ws/status/{project_id}")
async def websocket_status(websocket: WebSocket, project_id: str):
    """Real-time status updates via WebSocket"""
    await websocket.accept()
    
    try:
        while True:
            # Get latest status from cache/db
            status = await get_project_status(project_id)
            
            await websocket.send_json({
                "project_id": project_id,
                "stage": status.get("stage"),
                "progress_percent": status.get("progress"),
                "message": status.get("message"),
                "timestamp": datetime.now(timezone.utc).isoformat()
            })
            
            if status.get("status") in ["complete", "error"]:
                break
            
            await asyncio.sleep(2)  # Update every 2 seconds
    
    except Exception as e:
        await websocket.close(code=1000)

# ============================================================================
# 3. STAGE 1: PARSE EXCEL & EXTRACT PROJECT INFO
# ============================================================================

async def parse_excel_template(workbook) -> Dict:
    """
    Extract project info and component data from Excel template.
    Returns: {"project_info": {...}, "components": {...}}
    """
    
    try:
        # Sheet 1: PROJECT_INFO
        project_sheet = workbook["PROJECT_INFO"]
        
        project_info = {
            "project_name": project_sheet["B3"].value,
            "client_name": project_sheet["B4"].value,
            "client_email": project_sheet["B5"].value,
            "address": project_sheet["B6"].value,
            "city_state_zip": project_sheet["B7"].value,
            "jurisdiction": project_sheet["B8"].value,
            "scope": project_sheet["B9"].value,
            "special_conditions": project_sheet["B10"].value,
        }
        
        # Sheet 2: COMPONENTS_SCHEDULE (user-provided quantities)
        component_sheet = workbook["COMPONENTS_SCHEDULE"]
        
        excel_components = {}
        for row in component_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Component type
                excel_components[row[0]] = {
                    "quantity": row[1] or 0,
                    "unit": row[2] or "each",
                    "size_spec": row[3],
                    "location": row[4],
                    "notes": row[5],
                    "source": "excel",
                    "confidence": 1.0
                }
        
        return {
            "project_info": project_info,
            "excel_components": excel_components
        }
    
    except Exception as e:
        return {
            "project_info": None,
            "excel_components": {},
            "error": f"Failed to parse Excel: {str(e)}"
        }

# ============================================================================
# 4. STAGE 2: EXTRACT DATA FROM PDFs (Parallel Processing)
# ============================================================================

async def extract_from_pdfs_parallel(pdf_bytes_list: List[bytes]) -> Dict:
    """
    Process multiple PDFs in parallel using asyncio.
    Each PDF: PDF → Images → OCR → Table extraction → Component recognition
    """
    
    all_extracted = {
        "components": {},
        "drawings": {},
        "specs": {},
        "confidence": {}
    }
    
    # Process all PDFs concurrently
    tasks = [
        extract_single_pdf(pdf_bytes)
        for pdf_bytes in pdf_bytes_list
    ]
    
    results = await asyncio.gather(*tasks, return_exceptions=True)
    
    # Merge results from all PDFs
    for result in results:
        if isinstance(result, dict):
            # Merge component data
            for comp_type, data in result.get("components", {}).items():
                if comp_type not in all_extracted["components"]:
                    all_extracted["components"][comp_type] = data
                else:
                    # If component appeared in multiple PDFs, take average confidence
                    prev_conf = all_extracted["confidence"].get(comp_type, 1.0)
                    new_conf = data.get("confidence", 1.0)
                    all_extracted["confidence"][comp_type] = (prev_conf + new_conf) / 2
    
    return all_extracted

async def extract_single_pdf(pdf_bytes: bytes) -> Dict:
    """Extract components and specs from a single PDF"""
    
    try:
        # Step 1: Convert PDF to images
        images = await pdf_to_images(pdf_bytes)
        
        extracted = {
            "components": {},
            "drawings": {},
            "specs": {}
        }
        
        # Step 2: For each page, run OCR + vision models
        for page_num, image in enumerate(images):
            
            # Step 2a: Extract text via OCR
            text = await run_ocr(image)
            
            # Step 2b: Find tables in image (for component schedules)
            tables = await find_tables_in_image(image)
            
            # Step 2c: Parse each table for component data
            for table in tables:
                components = parse_component_table(table, text)
                extracted["components"].update(components)
            
            # Step 2d: Extract specs from drawings
            specs = extract_specs_from_image(image, text)
            extracted["specs"].update(specs)
        
        return extracted
    
    except Exception as e:
        return {
            "error": str(e),
            "components": {},
            "specs": {}
        }

async def pdf_to_images(pdf_bytes: bytes) -> List:
    """Convert PDF to list of images (one per page)"""
    
    pdf_reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
    images = []
    
    for page_num, page in enumerate(pdf_reader.pages):
        # Use pdfplumber for better extraction
        try:
            # This is simplified - in reality use pdf2image or pdfplumber
            image = await convert_pdf_page_to_image(page, page_num)
            images.append(image)
        except:
            pass
    
    return images

def parse_component_table(table: List[List], context_text: str) -> Dict:
    """
    Parse a table extracted from PDF.
    Table format: 
      Headers: ["Component Type", "Qty", "Size", "Spec"]
      Rows: [["Windows", "12", "3'x5'", "Vinyl"], ...]
    """
    
    components = {}
    
    if len(table) < 2:
        return components
    
    headers = [str(h).lower() for h in table[0]]
    
    # Find relevant columns
    qty_col = None
    type_col = None
    spec_col = None
    
    for i, header in enumerate(headers):
        if any(x in header for x in ["qty", "quantity", "count", "number"]):
            qty_col = i
        elif any(x in header for x in ["type", "component", "item", "description"]):
            type_col = i
        elif any(x in header for x in ["spec", "size", "model", "details"]):
            spec_col = i
    
    # If we found quantity column, parse rows
    if qty_col is not None:
        for row in table[1:]:
            if len(row) > qty_col:
                try:
                    qty = int(row[qty_col])
                    comp_type = str(row[type_col]) if type_col else "unknown"
                    spec = str(row[spec_col]) if spec_col else None
                    
                    if qty > 0:  # Only include non-zero quantities
                        components[comp_type] = {
                            "quantity": qty,
                            "spec": spec,
                            "source": "ai_extracted",
                            "confidence": 0.85  # Table extraction is pretty reliable
                        }
                except (ValueError, IndexError, TypeError):
                    continue
    
    return components

# ============================================================================
# 5. STAGE 3: MERGE DATA (Excel + AI Extraction)
# ============================================================================

def merge_component_data(
    excel_components: Dict,
    ai_extracted: Dict
) -> Dict:
    """
    Smart merge of Excel-provided vs AI-extracted component data.
    
    Priority:
    1. User provided in Excel (trust it, confidence 1.0)
    2. AI extracted from PDF (use if not in Excel)
    3. Combine if both exist (average confidence)
    """
    
    merged = {}
    
    # Start with Excel data (highest priority)
    merged.update(excel_components)
    
    # Add AI-extracted data
    for comp_type, ai_data in ai_extracted.get("components", {}).items():
        if comp_type in merged:
            # Component in both - Excel takes priority but note AI confidence
            merged[comp_type]["ai_confidence"] = ai_data.get("confidence", 0.7)
        else:
            # Only in AI extraction
            merged[comp_type] = ai_data
    
    return merged

# ============================================================================
# 6. STAGE 4: RUN COMPLIANCE ANALYSIS (Parallel)
# ============================================================================

async def run_compliance_analysis(
    components: Dict,
    jurisdiction: str,
    zip_code: str
) -> List[Dict]:
    """
    Run all compliance rules in parallel:
    - IRC 2018 checks
    - IECC 2015 checks
    - NEC 2017 checks
    - Jurisdiction amendments (Georgia, etc.)
    """
    
    # Get regional factors (needed for some rules)
    regional_factors = await get_regional_factors(zip_code)
    
    # Run all rule sets concurrently
    tasks = [
        check_irc_rules(components, regional_factors),
        check_iecc_rules(components, regional_factors),
        check_nec_rules(components, regional_factors),
        check_jurisdiction_amendments(components, jurisdiction, regional_factors)
    ]
    
    results = await asyncio.gather(*tasks, return_exceptions=True)
    
    # Flatten all findings
    all_findings = []
    for result in results:
        if isinstance(result, list):
            all_findings.extend(result)
    
    return all_findings

async def check_irc_rules(components: Dict, factors: Dict) -> List[Dict]:
    """Call real rules service with components"""
    
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(
                f"{RULES_SERVICE_URL}/check-irc",
                json={
                    "components": components,
                    "year": 2018,
                    "state": factors.get("state")
                }
            ) as resp:
                if resp.status == 200:
                    return await resp.json()
    except:
        pass
    
    return []

async def check_iecc_rules(components: Dict, factors: Dict) -> List[Dict]:
    """Call real rules service"""
    
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(
                f"{RULES_SERVICE_URL}/check-iecc",
                json={
                    "components": components,
                    "climate_zone": factors.get("climate_zone"),
                    "year": 2015
                }
            ) as resp:
                if resp.status == 200:
                    return await resp.json()
    except:
        pass
    
    return []

async def check_nec_rules(components: Dict, factors: Dict) -> List[Dict]:
    """Call real rules service"""
    
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(
                f"{RULES_SERVICE_URL}/check-nec",
                json={
                    "components": components,
                    "year": 2017
                }
            ) as resp:
                if resp.status == 200:
                    return await resp.json()
    except:
        pass
    
    return []

# ============================================================================
# 7. STAGE 5: GENERATE COST ESTIMATE (Parallel)
# ============================================================================

async def generate_cost_estimate(
    components: Dict,
    zip_code: str,
    jurisdiction: str
) -> Dict:
    """Generate detailed cost estimate"""
    
    # Get pricing data (regional factors)
    pricing = await get_regional_pricing(zip_code)
    
    estimate = {
        "line_items": [],
        "subtotal": 0,
        "labor_rate": pricing.get("labor_rate", 50),
        "material_factor": pricing.get("material_factor", 1.0),
        "regional_info": {
            "zip_code": zip_code,
            "region": pricing.get("region"),
            "county": pricing.get("county")
        }
    }
    
    # For each component, generate cost estimate
    for comp_type, comp_data in components.items():
        qty = comp_data.get("quantity", 0)
        
        # Look up unit cost from pricing database
        unit_cost = await get_unit_cost(comp_type, zip_code)
        
        line_item = {
            "component": comp_type,
            "quantity": qty,
            "unit_cost": unit_cost,
            "line_total": qty * unit_cost,
            "labor_hours": qty * get_labor_hours(comp_type),
            "material_cost": qty * unit_cost
        }
        
        estimate["line_items"].append(line_item)
        estimate["subtotal"] += line_item["line_total"]
    
    # Calculate O&P and contingency
    estimate["overhead_profit"] = estimate["subtotal"] * 0.35
    estimate["contingency"] = (estimate["subtotal"] + estimate["overhead_profit"]) * 0.10
    estimate["permit_allowance"] = 2500  # Typical GA permit
    
    estimate["grand_total"] = (
        estimate["subtotal"] +
        estimate["overhead_profit"] +
        estimate["contingency"] +
        estimate["permit_allowance"]
    )
    
    return estimate

# ============================================================================
# 8. STAGE 6: GENERATE OUTPUTS (Excel, PDF, CSV)
# ============================================================================

async def generate_outputs(
    project_id: str,
    project_info: Dict,
    components: Dict,
    findings: List,
    estimate: Dict
) -> Dict:
    """Generate all downloadable files"""
    
    outputs = {}
    
    # Output 1: Updated Excel with findings + estimate
    excel_file = generate_populated_excel(
        project_info, components, findings, estimate
    )
    outputs["excel"] = excel_file
    
    # Output 2: Professional PDF proposal
    pdf_proposal = generate_pdf_proposal(
        project_info, components, findings, estimate
    )
    outputs["proposal_pdf"] = pdf_proposal
    
    # Output 3: Xactimate CSV export
    xactimate_csv = generate_xactimate_csv(components, estimate)
    outputs["xactimate_csv"] = xactimate_csv
    
    # Output 4: Compliance report
    compliance_report = generate_compliance_report(findings)
    outputs["compliance_report"] = compliance_report
    
    return outputs

def generate_populated_excel(
    project_info: Dict,
    components: Dict,
    findings: List,
    estimate: Dict
) -> bytes:
    """Create Excel with auto-populated findings and estimate"""
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: Project Info
    ws = wb.active
    ws.title = "PROJECT_INFO"
    ws["A1"] = "Project Information"
    ws["B3"] = project_info.get("project_name")
    # ... populate all fields
    
    # Sheet 2: Findings
    ws = wb.create_sheet("FINDINGS_OUTPUT")
    ws["A1"] = "Code Compliance Findings"
    ws["A2"] = "Severity"
    ws["B2"] = "Code"
    ws["C2"] = "Finding"
    ws["D2"] = "Recommendation"
    
    row = 3
    for finding in findings:
        ws[f"A{row}"] = finding.get("severity")
        ws[f"B{row}"] = finding.get("code")
        ws[f"C{row}"] = finding.get("description")
        ws[f"D{row}"] = finding.get("recommendation")
        
        # Color code severity
        if finding.get("severity") == "RED":
            ws[f"A{row}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif finding.get("severity") == "ORANGE":
            ws[f"A{row}"].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        
        row += 1
    
    # Sheet 3: Estimate
    ws = wb.create_sheet("ESTIMATE_OUTPUT")
    ws["A1"] = "Cost Estimate"
    ws["A2"] = "Line Item"
    ws["B2"] = "Qty"
    ws["C2"] = "Unit Cost"
    ws["D2"] = "Total"
    
    row = 3
    for line_item in estimate.get("line_items", []):
        ws[f"A{row}"] = line_item.get("component")
        ws[f"B{row}"] = line_item.get("quantity")
        ws[f"C{row}"] = f"${line_item.get('unit_cost'):.2f}"
        ws[f"D{row}"] = f"${line_item.get('line_total'):.2f}"
        row += 1
    
    # Totals row
    ws[f"A{row}"] = "SUBTOTAL"
    ws[f"D{row}"] = f"${estimate.get('subtotal'):.2f}"
    ws[f"D{row}"].font = Font(bold=True)
    
    row += 1
    ws[f"A{row}"] = "Overhead & Profit (35%)"
    ws[f"D{row}"] = f"${estimate.get('overhead_profit'):.2f}"
    
    row += 1
    ws[f"A{row}"] = "Contingency (10%)"
    ws[f"D{row}"] = f"${estimate.get('contingency'):.2f}"
    
    row += 1
    ws[f"A{row}"] = "Permit Allowance"
    ws[f"D{row}"] = f"${estimate.get('permit_allowance'):.2f}"
    
    row += 1
    ws[f"A{row}"] = "GRAND TOTAL"
    ws[f"D{row}"] = f"${estimate.get('grand_total'):.2f}"
    ws[f"D{row}"].font = Font(bold=True, size=14)
    
    # Save to bytes
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    return excel_bytes.getvalue()

# ============================================================================
# 9. MAIN ORCHESTRATION
# ============================================================================

async def run_full_analysis(
    project_id: str,
    excel_workbook,
    pdf_bytes_list: List[bytes]
):
    """
    Complete analysis pipeline orchestration.
    Runs Stages 1-6 in optimal sequence.
    """
    
    try:
        # STAGE 1: Parse Excel
        await update_status(project_id, "stage_1_parse_excel", 10, "Parsing Excel template...")
        
        excel_parsed = await parse_excel_template(excel_workbook)
        project_info = excel_parsed.get("project_info")
        excel_components = excel_parsed.get("excel_components", {})
        
        # STAGE 2: Extract from PDFs (Parallel)
        await update_status(project_id, "stage_2_extract_pdfs", 25, "Extracting data from PDFs...")
        
        ai_extracted = await extract_from_pdfs_parallel(pdf_bytes_list)
        
        # STAGE 3: Merge data
        await update_status(project_id, "stage_3_merge", 35, "Merging Excel + AI data...")
        
        merged_components = merge_component_data(excel_components, ai_extracted)
        
        # STAGE 4: Run compliance analysis (Parallel)
        await update_status(project_id, "stage_4_compliance", 50, "Checking code compliance...")
        
        findings = await run_compliance_analysis(
            merged_components,
            project_info.get("jurisdiction"),
            project_info.get("zip_code")
        )
        
        # STAGE 5: Generate estimate (Parallel)
        await update_status(project_id, "stage_5_pricing", 65, "Calculating costs...")
        
        estimate = await generate_cost_estimate(
            merged_components,
            project_info.get("zip_code"),
            project_info.get("jurisdiction")
        )
        
        # STAGE 6: Generate outputs
        await update_status(project_id, "stage_6_outputs", 80, "Generating documents...")
        
        outputs = await generate_outputs(
            project_id,
            project_info,
            merged_components,
            findings,
            estimate
        )
        
        # Save results
        await save_analysis_results(
            project_id,
            {
                "project_info": project_info,
                "components": merged_components,
                "findings": findings,
                "estimate": estimate,
                "outputs": outputs,
                "status": "complete"
            }
        )
        
        # STAGE 7: Notify user
        await update_status(project_id, "complete", 100, "Done! Check your email.")
        
        await send_notification_email(
            project_info.get("client_email"),
            project_id,
            project_info.get("project_name"),
            estimate.get("grand_total")
        )
    
    except Exception as e:
        await update_status(project_id, "error", 0, f"Analysis failed: {str(e)}")
        await send_error_email(project_id, str(e))

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

async def get_regional_factors(zip_code: str) -> Dict:
    """Look up regional data by ZIP code"""
    # In production: query database
    return {
        "state": "GA",
        "county": "Madison",
        "climate_zone": "4A",
        "labor_rate": 45,
        "material_factor": 1.05
    }

async def get_regional_pricing(zip_code: str) -> Dict:
    """Get pricing data by ZIP code"""
    # In production: query pricing database
    return {
        "region": "Northeast GA",
        "labor_rate": 45,
        "material_factor": 1.05
    }

async def get_unit_cost(component_type: str, zip_code: str) -> float:
    """Get unit cost for component type and location"""
    # In production: query pricing database
    UNIT_COSTS = {
        "windows": 250,
        "doors": 350,
        "framing": 1.25,  # per board foot
        "roofing": 4.50,  # per SF
        "drywall": 0.95,  # per SF
    }
    return UNIT_COSTS.get(component_type, 100)

def get_labor_hours(component_type: str) -> float:
    """Get labor hours per unit"""
    LABOR_HOURS = {
        "windows": 2.0,
        "doors": 1.5,
        "framing": 0.01,  # per BF
        "roofing": 0.03,  # per SF
    }
    return LABOR_HOURS.get(component_type, 1.0)

async def update_status(project_id: str, stage: str, progress: int, message: str):
    """Update project status in cache/DB"""
    # In production: save to Redis or DB
    pass

async def save_analysis_results(project_id: str, results: Dict):
    """Save analysis results to database"""
    # In production: save to PostgreSQL
    pass

async def send_notification_email(email: str, project_id: str, project_name: str, total: float):
    """Send completion email to user"""
    # In production: use SendGrid or similar
    pass

# ============================================================================
# Launch
# ============================================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8005)
```

---

## Time Breakdown: 5-10 Minutes Per Project

```
Stage 1: Parse Excel           ~ 30 seconds
Stage 2: Extract PDFs (parallel)  ~ 2-3 minutes
Stage 3: Merge data             ~ 30 seconds
Stage 4: Compliance checking (parallel) ~ 1-2 minutes
Stage 5: Cost estimation        ~ 1 minute
Stage 6: Generate outputs       ~ 30 seconds
Stage 7: Send email             ~ 30 seconds
─────────────────────────────────────────
TOTAL:                          ~ 5-10 minutes ✅
```

Compared to manual process: **10-14 hours → 5-10 minutes = 60x faster**

---
